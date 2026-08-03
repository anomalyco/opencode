from hashlib import sha256
from json import dump
from pathlib import Path
from sys import argv, stdout

from openpyxl import load_workbook


path = Path(argv[1]).resolve(strict=True)
workbook = load_workbook(path, read_only=True, data_only=True)
sheet = workbook["清洗结果"]
stdout.reconfigure(encoding="utf-8", errors="strict")
dump(
    {
        "fileName": path.name,
        "sha256": sha256(path.read_bytes()).hexdigest(),
        "rows": [[cell for cell in row] for row in sheet.iter_rows(values_only=True)],
    },
    stdout,
    ensure_ascii=False,
    default=str,
)
